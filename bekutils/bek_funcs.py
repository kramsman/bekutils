""" small functions that can be imported and used in other programs.
Taken from ROVCleaver
"""

# TODO rename bekutils
# TODO what to do with loggers?


# import ast
# import collections
# import datetime
# import importlib
# import inspect
# import itertools
# import math  # for ceil function
# import os
# import pathlib
# import re
# import shutil  # to copy file from other dirs
# import webbrowser
# from datetime import datetime
# from itertools import islice  # to skip 1st row of iterated spreadsheet
# from pathlib import Path
# from tkinter import Tk  # from tkinter import Tk for Python 3.x
# from typing import Union

# from tkinter.filedialog import askopenfilename

# import numpy as np
# import pandas as pd
# os.environ["APPDATA"] = ""
# from pandasgui import show
# import pymsgbox
# from openpyxl import load_workbook
# from openpyxl.styles import Font
# import PySimpleGUI as sg
# import sys
# from openpyxl.utils.cell import coordinate_from_string
# from openpyxl.utils.cell import column_index_from_string
# import json
# from loguru import logger

log_level = "DEBUG"  # used for log file; screen set to INFO. TRACE, DEBUG, INFO, WARNING, ERROR


def bek_write_excel(df, sheet_name, startrow, cell_infos = None,):
    """
    Write df to an excel file with the same name as the py file suffixed with xlsx in the current directory.
    Autosizes before titles to avoid ##### in cells.
    Args:
        df (): dataframe being written out
        sheet_name ():
        startrow (): the row where the df rows will begin to be written
        cell_infos (): list of cell attributes used to format specified as dictionaries. Rows and cols 1 based
        numerics. 'value's are replaced as text; font is passed as eval of cell value (note in function):
            {'row':1,'col':1, 'cell_attr':"value", 'cell_value':'Summary Report'},
            {'row':1,'col':1, 'cell_attr':"font", 'cell_value':'Font(b=True, size=20)'},
    """

    from pathlib import Path
    import pandas as pd
    from openpyxl.styles import Font
    from bekutils import autosize_xls_cols
    from bekutils import exe_path

    op_file = f"{Path(__file__).stem}.xlsx"
    writer = pd.ExcelWriter(op_file)

    df.to_excel(writer, sheet_name = sheet_name, startrow = startrow)
    wb = writer.book
    for sh in wb.worksheets:
        autosize_xls_cols(sh)

    if cell_infos:
        for sh in wb.worksheets:
            for cell_info in cell_infos:
                if cell_info['cell_attr'] == 'font':
                    # use eval because needs to be like ft1 = Font(name='Arial', size=14).  might be able to use
                    # another setattr but not ready to try now
                    setattr(sh.cell(row=cell_info['row'], column=cell_info['col']), cell_info['cell_attr'],
                            eval(cell_info['cell_value']))
                else:
                    setattr(sh.cell(row=cell_info['row'], column=cell_info['col']), cell_info['cell_attr'],
                            cell_info['cell_value'])

    writer.close()

def exe_file():
    """ return the file of location where exe is running """

    import sys
    from pathlib import Path

    import __main__

    # determine if application is running as a script file or frozen exe
    if getattr(sys, 'frozen', False):
        exe_file = Path(sys.executable)
    elif __file__:
        # exe_file = Path(__file__).parents[0]
        exe_file = Path(__main__.__file__)
    else:
        exe_file = None

    return exe_file


def exe_path():
    """ return the path of location where exe is running """

    from bekutils import exe_file
    # from pathlib import Path
    #
    # import __main__
    # import os

    exe_path = exe_file().parents[0]

    return exe_path


def setup_loguru(log_level_std='INFO', log_level_log='INFO', log_path=None, log_mode='w'):
    """ set log file path to location based on whether setup file is used or not.  If not, use EXE.
        using loguru could put in downloads (next line)
        but I'm locating it once we know what section
        is running, either with the setup file or the EXE if no setup, eg updating zip file.
        logfile = pathlib.Path.home() / "Downloads" / (Path(__file__).name + ".log")
    """

    from pathlib import Path
    from loguru import logger
    import os
    import sys
    from bekutils import exe_path
    from bekutils import exe_file

    # LOG_LEVEL_LOG = "TRACE"  # used for log file; screen set to INFO. TRACE, DEBUG, INFO, WARNING, ERROR
    # LOG_LEVEL_STD = "DEBUG"  # used for log file; screen set to INFO. TRACE, DEBUG, INFO, WARNING, ERROR

    if log_path is None:
        log_path = exe_path()

        # # determine if application is running as a script file or frozen exe
        # if getattr(sys, 'frozen', False):
        #     log_path = Path(sys.executable).parents[0]
        # elif __file__:
        #     # root_path = os.path.dirname(__file__)
        #     log_path = Path(__file__).parents[0]
        # else:
        #     log_path = None
    logger.debug(f"({log_path=}")

    logger.remove(0)
    # logfile = f"{exe_file().stem}.xlsx"
    logfile = exe_file().with_suffix(".log")
    try:
        os.remove(logfile)
    except Exception as e:
        logger.exception(e)
        pass

    logger.add(open(logfile, log_mode), level=log_level_log, backtrace=True, diagnose=True)
    logger.add(sys.stdout, level=log_level_std, backtrace=True, diagnose=True)

    return logger


def exit_yes_no(msg, title=None, display_exiting=False):
    """ makes this choice to continue one line"""

    import pymsgbox
    from loguru import logger

    if not title:
        title = "Exit?"
    choice = pymsgbox.confirm(msg, title, ['Yes', 'No'])
    if choice == "No":
        if display_exiting:
            pymsgbox.alert("Exiting", "Alert")
        logger.debug("here'")
        exit()


def exit_yes(msg: str, title: str = None, *, errmsg: str = None, raise_err: bool=False) -> None:
    """ exits program after giving user a popup window and raising an error. """

    import pymsgbox
    from loguru import logger

    msg = msg + "\n\n\nExiting."
    if not errmsg:
        errmsg = msg.replace("\n", " ")  # dont fill the console with linefeeds
    if not title:
        title = "** Exiting Program **"
    logger.debug("in 'exit_yes'")
    pymsgbox.alert(msg, title)
    if raise_err:
        raise Exception(errmsg)


def is_number(s: str) -> bool:
    """  Used as check, particularly before trying to set zip to numeric for lookup.
    expects param to be a string to trap all types of data.   """

    import numpy as np
    from loguru import logger

    logger.info("in is_number")

    try:
        if np.isnan(s):  # this is needed- np.nan are int which are numbers
        # if s == np.nan:  # this is needed- np.nan are int which are numbers
            return False
    except TypeError:
        pass
    if s is None:
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False



def clean_field(fld, case_convert='lower'):
    """
    returns a string in lower, strip, no space, no -, no ., no '
    can be used with dataframe like IP['clean2'] = IP['B'].apply(clean_field, convert_case='keep')
    1/28/23 added optional parameter convert_case defaulting to lower, as was done before, but allowing 'upper' or
    'keep'.
    """

    #TODO: pass characters to be removed as string

    from loguru import logger

    return_fld = str(fld).strip().replace(" ", "").replace("'", "").replace(".", "").replace("-", "")
    if case_convert == 'lower':
        return_fld = return_fld.lower()
    elif case_convert == 'upper':
        return_fld = return_fld.upper()
    elif case_convert == 'keep':
        pass
    else:
        exit_yes(f"wrong value fed to clean_field parameter case_convert, '{case_convert}' - exiting")
    return return_fld


def autosize_xls_cols(ws):
    """ BEKs routine that works on the wks rather than df.  Datetime format set to width of 10. """

    from loguru import logger

    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                if cell.data_type == 'd':
                    date_width = 10
                else:
                    date_width = len(str(cell.value))
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), date_width))

    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 1


def bad_file_exit(file, msg=None, raise_err=False):
    """ checks for file existence and exits if not found"""

    from loguru import logger

    if msg is None:
        msg = f"File:\n\n'{file}'\n\ndoes not exist."
    if not file.expanduser().exists():
        logger.debug("here")
        exit_yes(msg, raise_err=raise_err)


def bad_path_exit(path, msg=None, raise_err=False):
    """ checks for directory existence and exits if not found"""

    from loguru import logger

    if msg is None:
        msg = f"Directory:\n\n'{path}'\n\ndoes not exist."
    # if not Path(os.path.expanduser(path)).exists():  # need expanduser for ~; only os works (not pathlib)
    if not path.expanduser().exists():
        # pymsgbox.alert(msg, "** Exiting via bad_path_exit **")
        # # FIXME: close TKINTER window here.  https://stackoverflow.com/questions/8009176/function-to-close-the-window-in-tkinter
        # exit()
        logger.debug("here")
        exit_yes(msg, raise_err=raise_err)


def bad_path_create(path, msg=None):
    """ checks for directory existence and creates if not found"""

    import os
    import pymsgbox
    from loguru import logger

    if msg is None:
        msg = ("Directory:\n\n" + str(path) + "\n\ndoes not exist.  Creating." +
               "\n\nCalled from " + calling_func(level=2))
    if not os.path.isdir(path):
        logger.debug("here")
        pymsgbox.alert(msg, "Adding Directory via bad_path_create")
        os.makedirs(path)


def calling_func(level=0):
    """ returns the various levels of calling function.  0 is current, 1 is caller of current, etc """

    import inspect
    from loguru import logger

    try:
        func = f"'{inspect.stack()[level][3]}', line #: {inspect.stack()[level][2]}"
    except Exception as e:
        logger.exception(e)
        func = f"** error ** inspect level too deep: {str(level)} called from {inspect.stack()[level][3]}"
    return func


def read_file_to_df(file_with_path, **param_dict):
    """reads either xlsx or csv into a dataframe using parms passed in dictionary. Non-applicable parms
    are skipped."""

    import inspect
    from pathlib import PurePath
    import pandas as pd
    from loguru import logger

    logger.info(f"reading file to dataframe '{file_with_path.stem}'")

    if PurePath(file_with_path).suffix.lower() == '.xlsx':
        filtered_dict = {k: v for k, v in param_dict.items()
                         if k in [p.name for p in inspect.signature(pd.read_excel).parameters.values()]}
        df_temp = pd.read_excel(file_with_path, **filtered_dict)
    elif file_with_path.suffix.lower() == '.csv':
        filtered_dict = {k: v for k, v in param_dict.items()
                         if k in [p.name for p in inspect.signature(pd.read_csv).parameters.values()]}
        df_temp = pd.read_csv(file_with_path, **filtered_dict)
    else:
        df_temp = None
        logger.debug('here')
        exit_yes((f"Bad file type on input - not xlsx or csv."
                  f"\n\nFile: '{file_with_path}'"
                  ))
    return df_temp



def find_header_row_in_file(file_with_path, header_string, header_col, sheet_name=None):
    """ identifies row with header by searching for header_string in header_col.  Used to skip blank and rows with titles.

    Parameters
    ----------
    file_with_path : input file being read in, csv or xlsx?
    header_string : string identifying header row, like 'pdiid'
    header_col : alpha col to search for string, 'B' 'AA'
    sheet_name : sheet name in input file in case multiple
    """

    from openpyxl.utils.cell import coordinate_from_string
    from openpyxl.utils.cell import column_index_from_string
    from bekutils import read_file_to_df

    if sheet_name is None:
        sheet_name = 0
    header_row = None

    # If the header identifying field is not in the first 30 rows assume something is wrong in the file
    df_temp = read_file_to_df(file_with_path, **{'header': None, 'sheet_name': sheet_name, 'nrows': 30,
                                                 'keep_default_na': True, 'dtype': str})

    excel_col_num = column_index_from_string(coordinate_from_string(header_col + '1')[0])  # -1 to 0 index

    for row in df_temp.itertuples():
        if type(row[excel_col_num]) == str:  # cell in input being checked is ok, otherwise blank cells/None cause
            # problems in compare
            if row[excel_col_num].strip().lower() == header_string.strip().lower():
                header_row = row.Index
                break
    if header_row is None:
        exit_yes((f"File may be bad.\n\nThe header check string '{header_string}' "
                  f"was not found in column '{header_col}' "
                  "in the first 30 lines of input file:"
                  f"\n\n'{file_with_path}'"
                  ))
    return header_row


def check_ws_headers(ws, vals):
    """
    Check list of (cell, val) tuples representing header labels in ws_to_chk and error if val not found in cell.
    eg vals = [('A1', 'use'), ('B1', 'fromFilePath'), ('C1', 'fromfilename'), ....]
    """

    def chk_header_vals(ws_to_chk, cell, val):
        """ error if val not found in wks cell. """
        if str(ws_to_chk[cell].value).strip().lower() != str(val).lower():
            exit_yes((f"Column heading '{cell}' on Setup sheet '{ws_to_chk.title}' not equal to literal '{val}'."
                      f"\n\nIt is '{str(ws_to_chk[cell].value)}'."),
                     )

    for pairs in vals:
        chk_header_vals(ws, pairs[0], pairs[1])

# TODO Add in check_fie+headers like above with csv

def text_box(txt, title='', box_title="", buttons=None):
    """ Display text block with lines separated by \n and choice of buttons at bottom.

    Parameters
    ----------
    box_title :
    title :
    txt :
    buttons :

    """

    import PySimpleGUI as sg

    # window = sg.Window('Virus Simulation', layout, background_color='hex_color_code')

    if buttons is None:
        buttons = ["OK", "Exit"]

    col_factor = 3  # to scale window equally
    row_factor = 30  # to scale window equally
    max_cols = len(max(txt.split("\n"), key=len)) * col_factor
    cols = max_cols
    # v_scroll = False
    col_limit = 80 * col_factor
    col_min = 50 * col_factor
    if cols > col_limit:
        # v_scroll = True
        cols = col_limit
    elif cols < col_min:
        cols = col_min

    noscroll = True
    row_limit = 80
    row_min = 6
    # max_rows = len(txt.split("\n"))
    # rows = max_rows
    rows = len(txt.split("\n"))
    if rows > row_limit:
        noscroll = False
        rows = row_limit
    elif rows < row_min:
        rows = row_min
    #horizontal_scroll=h_scroll,
    # sg.theme('SystemDefault1')
    sg.theme('Default1')
    layout = [
        [sg.Text(title, font=("Arial", 18))],
        [sg.Multiline(txt, autoscroll=False, expand_x=True, no_scrollbar=noscroll,
                      expand_y=True, enable_events=True)],
        [sg.Button(text) for text in buttons],
    ]

    event, values = sg.Window(box_title, layout, titlebar_font=("Arial", 20), font=("Arial", 14),
                              use_custom_titlebar=True, size=(600, rows*row_factor), disable_close=True,
                              resizable=True, grab_anywhere=True).read(close=True)
    if event is not None:
        event = event.lower()
    return event


def get_dir_name(box_title, title2, initial_dir):
    """ show an "Open" dialog box and return the selected directory. Replaced askdirectory with PySimpleGUI
    :param title2:
    :type title2:
    """

    import os
    import PySimpleGUI as sg
    from pathlib import Path
    from loguru import logger

    logger.debug('in get_dir_name')

    layout = [
        [sg.Text(title2, font=("Arial", 18))],
        [
         sg.Input(key="-IN-", expand_x=True),
         sg.FolderBrowse(initial_folder=os.path.expanduser(initial_dir))
         ],
        [sg.Button("Choose")],
    ]

    # event, values = sg.Window(heading_in_box, layout, size=(600, 100)).read(close=True)
    event, values = sg.Window(box_title, layout, titlebar_font=("Arial", 20), font=("Arial", 14),
                              size=(1000, 150), use_custom_titlebar=True).read(close=True)

    dir_name = values['-IN-']
    if dir_name == "":
        exit_yes("No directory name chosen")

    return Path(dir_name).expanduser()


def get_file_name(box_title, title2, initial_dir):
    """ show an "Open" dialog box and return the selected file name. Replaced askopenfilename with pyeasygui
    :param title2: heading of the box
    :type title2: text next to input field
    """

    import os
    import PySimpleGUI as sg
    from pathlib import Path
    from loguru import logger

    logger.debug('in get_file_name')
    # "Select Sincere address export file 'all-parent-campaign-requests-yyyy-mm-dd.csv'"
    layout = [
        [sg.Text(title2, font=("Arial", 18))],
        [
         sg.Input(key="-IN-", expand_x=True),
         sg.FileBrowse(initial_folder=os.path.expanduser(initial_dir))
         ],
        [sg.Button("Choose")],
    ]

    # event, values = sg.Window(heading_in_box, layout, size=(600, 100)).read(close=True)
    event, values = sg.Window(box_title, layout, titlebar_font=("Arial", 20), font=("Arial", 14),
                              size=(1000, 150), use_custom_titlebar=True).read(close=True)
    # sg.Window.close()

    file_name = values['-IN-']
    if file_name == "":
        exit_yes("No file name chosen")

    return Path(file_name).expanduser()


def convert_bool(bool_val):
    """ bool('FALSE') return True so need better """
    if isinstance(bool_val, bool):
        return_val = bool_val
    else:
        if bool_val is None or bool_val.lower() not in ['true', 'false']:
            raise ValueError('only allowable booleans are any case of true and false.  0/1 could be added to '
                             'convert_bool code')
        elif bool_val.lower() == 'true':
            return_val = True
        else:
            return_val = False
    return return_val


if __name__ == '__main__':

    fff = exe_file()
    ppp = exe_path()

    print(f"{clean_field('a-1-t')}")

    import numpy as np

    print(f"{is_number(np.nan)}")
    print(f"{is_number(1)}")
    print(f"{is_number('1')}")
    print(f"{is_number('a')}")
