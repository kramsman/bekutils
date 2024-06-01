""" show how path attribute could be added to and used with openpyxl """

from pathlib import Path
from openpyxl import load_workbook

def load_workbook_w_filepath(file, *args, **kwargs):
    """ openpyxl load_workbook that adds a filepath attribute

    Args:
        file (): PosixPath path and filename of workbook
    """
    from openpyxl import load_workbook
    wb = load_workbook(file, *args, **kwargs)
    wb.filepath = file
    return wb


if __name__ == '__main__':
    filenm = "~/Documents/Test Workbook.xlsx"
    f = Path(filenm).expanduser()

    # wb = load_workbook(filename=f, data_only=True)
    # print(f"{wb.path=}")
    # print(f"{wb.archive=}")
    # print()

    # wb_w_filepath = load_workbook_w_filepath(f)
    wb_w_filepath = load_workbook_w_filepath(f, 'ttt', data_only=True)
    print(f"{wb_w_filepath.path=}")
    print(f"{wb_w_filepath.filepath=}")
    print(f"{wb_w_filepath.filepath.name=}")


a=1
